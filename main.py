import datetime
import time
import unittest
import ddt

import openpyxl as openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By


def read_excel():
    # 打开Excel表格
    data1 = openpyxl.load_workbook(r'D:\fuxi\common\login_data.xlsx')
    # 获取Excel表格里面的下标login
    table = data1['login']
    # 循环出Excel表格里面的数据
    a = []
    for rows in range(2, table.max_row + 1):
        b = []
        for cols in range(2, table.max_column + 1):
            b.append(table.cell(rows, cols).value)
        a.append(b)
    return a


time1 = datetime.datetime.now()


@ddt.ddt
class TestCase(unittest.TestCase):
    # 引用read_excel方法 获取表格里面的数据

    def setUp(self) -> None:

        print("用例开始执行", time1, )

    @ddt.data(*read_excel())
    @ddt.unpack
    def test_login_01(self, a, b, c):
        driver = webdriver.Chrome()
        driver.get('http://localhost/')
        driver.find_element(By.NAME, "username").send_keys(a)
        driver.find_element(By.NAME, "password").send_keys(b)
        driver.find_element(By.NAME, "Submit").click()
        time.sleep(1)
        denglu = driver.find_element(By.XPATH, '//div[@align="center"]/b').text
        if denglu == '登录成功!':
            self.assertEqual(denglu, "登录成功!")
            self.assertNotEqual(denglu, "登录成功!")
            self.assertIn(denglu, "登录成功!")
            self.assertTrue(denglu, "登录成功!")
            self.assertFalse(denglu, "登录成功!")
        elif denglu == '您的用户名或密码有误!':
            self.assertEqual(denglu, '您的用户名或密码有误!')
        else:
            print("用例执行失败")

        time.sleep(2)
        driver.quit()

    def tearDown(self) -> None:
        print("用例执行结束", time1, )


if __name__ == '__main__':
    suiet = unittest.TestSuite()
    suiet.addTest(TestCase("test_login_01"))

    runer = unittest.TextTestRunner()
    runer.run(suiet)
