import time
import openpyxl
from selenium.webdriver.common.by import By
from base.base_page import BasePage


def read_excel():
    # 打开Excel表格
    data1 = openpyxl.load_workbook(r'D:\py_project\chicv\data\login_data.xlsx')
    # 获取Excel表格里面的下标login
    table = data1['login']
    # 循环出Excel表格里面的数据
    a = []
    for rows in range(2, table.max_row + 1):
        b = []
        for cols in range(2, table.max_column + 1):
            b.append(table.cell(rows, cols).value)
        a.append(b)
    return a



class LoginPage(BasePage):
    # 页面的元素
    username_loc = (By.XPATH, '//div[@class="inp-item"]//input[@name="email"]')
    password_loc = (By.XPATH, '//div[@class="inp-item inp-password"]//input[@name="password"]')
    submit_loc = (By.XPATH, '//li[@class="login-item item"]//button[@type="submit"]')

    def login_ecshop(self, username='admin', password='admin', loginauth='admin'):
        self.send_keys(LoginPage.username_loc, username)
        self.send_keys(LoginPage.password_loc, password)
        self.click(LoginPage.submit_loc)
        time.sleep(3)

    # 断言
    def get_except_result(self):
        time.sleep(2)
        return self.title()
